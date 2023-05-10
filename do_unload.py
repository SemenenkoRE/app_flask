
class UnloadQuery:

    """ Объект получает команду sql и выполняет выгрузку записей за период """

    __connection = None
    __cursor = None
    __log = ''
    __directory = ''

    def __init__(self, query_unload):

        self.__query_unload = query_unload
        self.__create_tb()

    def __do_connection(self):

        """ """

        import bd_login_passw as d

        name_user = d.name_user
        name_server = d.name_server
        password = d.password
        name_db = d.name_db

        import pymssql

        self.__connection = pymssql.connect(name_server, name_user, password, name_db)
        self.__cursor = self.__connection.cursor()

    def __create_tb(self):

        self.__do_connection()
        self.__cursor.execute(self.__query_unload)
        self.__fill_tb()
        # self.__do_commit()
        self.__close_connect()

    def __fill_tb(self):

        from openpyxl import load_workbook

        wb = load_workbook(self.set_directory() + '/files/base1.xlsx')

        sheet = wb['unload_table']

        i = 2

        for item in self.__cursor:
            sheet.cell(row=i, column=1).value = item[1]
            sheet.cell(row=i, column=2).value = item[2]
            sheet.cell(row=i, column=3).value = item[3]
            i += 1

        wb.save(self.set_directory() + '/files/report1.xlsx')

    @staticmethod
    def set_directory():

        """  """

        import pathlib as p

        return str(p.PurePath(__file__).parent)

    def __set_interior(self, sheet, i):

        """ """

        from openpyxl.styles.numbers import BUILTIN_FORMATS

        sheet[f'A2:A{i}'].number_format = 'DD.MM.YYYY'
        sheet[f'B2:B{i}'].number_format = BUILTIN_FORMATS[1]
        sheet.freeze_panes = "B1"

    def __close_connect(self):

        """ """

        self.__connection.close()


if __name__ == '__main__':

    x = UnloadQuery("SELECT * FROM sales")
