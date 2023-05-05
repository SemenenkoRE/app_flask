import datetime

class MakeBd:

    __connection = None
    __cursor = None
    __log = ''

    COMM_CR_TB = 'DROP TABLE IF EXISTS sales CREATE TABLE sales (id INT NOT NULL IDENTITY, dat DATE, article INT, weight INT);'

    def __do_connection(self):

        """ """

        name_user = 'sa'
        name_server = 'localhost'
        password = 'ert!@#ERT'
        name_db = 'db_sales'

        import pymssql

        self.__connection = pymssql.connect(name_server, name_user, password, name_db)
        self.__cursor = self.__connection.cursor()

    def create_tb(self):

        self.__do_connection()
        self.__cursor.execute(self.COMM_CR_TB)
        self.__do_commit()
        self.__close_connect()

    def fill_tb(self):

        from openpyxl import load_workbook

        wb = load_workbook(self.set_directory() + '/files/data.xlsx')

        sheet = wb['sales']

        self.__do_connection()
        i = 2

        while sheet.cell(row=i, column=1).value != '':

            try:

                content_command = self.form_command(datetime.datetime.date((sheet.cell(row=i, column=1).value)),
                                                    sheet.cell(row=i, column=2).value,
                                                    sheet.cell(row=i, column=3).value)
                self.__insert_row(content_command)
                self.__do_commit()

                i += 1

            except Exception as exc:

                pass

                self.__log = self.__log + f'{exc}; номер строки в таблице: {i} \n'

        self.__close_connect()

    @staticmethod
    def form_command(date_n, article, weight):
        return f"INSERT INTO sales VALUES ('{date_n}', {article}, {weight});"

    @staticmethod
    def set_directory():

        """  """

        import pathlib as p

        return str(p.PurePath(__file__).parent)

    def __insert_row(self, content_command):

        """ """
        self.__cursor.execute(content_command)
        self.__do_commit()

    def __do_commit(self):

        """ """

        self.__connection.commit()

    def __close_connect(self):

        """ """

        self.__connection.close()


if __name__ == '__main__':

    x = MakeBd()
    x.create_tb()
    x.fill_tb()

    pass

